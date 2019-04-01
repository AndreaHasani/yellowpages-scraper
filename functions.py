from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import Font, Alignment
import csv
import os
import random
from itertools import islice


dir_path = os.path.dirname(os.path.realpath(__file__))


def createExcelFile(data):
    wb = Workbook()
    ws = wb.active

    totalCol = 1
    for vertical in data:
        tableCol = totalCol
        for line in vertical:
            if type(line) != list:
                ws.merge_cells("A{0}:H{0}".format(tableCol))
                header = ws["A%s" % tableCol]
                header.font = Font(size=12, name='Calibri')
                header.alignment = Alignment(
                    horizontal="center", shrink_to_fit=False)
                ws["A%s" % tableCol] = line
                ws.append(["Name", "Website", "Facebook URL", "Email",
                           "Telephone Number", "City", "State", "Zip"])
                tableCol += 2
            else:
                ws.append(line)

        # Range A to H
        totalCol += len(vertical)
        ref = "A{0}:H{1}".format(tableCol - 1, totalCol)
        tab = Table(displayName="Table%s" % totalCol, ref=ref)

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        totalCol += 2

    wb.save("table1.xlsx")


def createCsv(data, vertical, location):
    """
    Write data to csv files

    data = [[row][row]]
    """
    locs, verts = readFiles()
    path = "./files/{}/{}.csv".format(vertical, location)
    print("Writting to %s" % path)

    with open(path, "a") as csv_file:
        csv_writter = csv.writer(csv_file, delimiter=",")
        for col in data:
            csv_writter.writerow(col)


def readFiles():
    """    Reads files and appends them to lists    """
    loc = []
    vert = []
    with open("locations.txt") as f:
        for line in f.readlines():
            loc.append(line.strip())

    with open("verticals.txt") as f:
        for line in f.readlines():
            vert.append(line.strip())

    return loc, vert


def createFolders():
    """    Create folders from verticals    """
    locations, verticals = readFiles()
    for vert in verticals:
        os.makedirs("./files/%s" % vert, exist_ok=True)


def processLines(lines):
    output = []
    for line in lines:
        newline = ["" if x.lower() == "empty" else x.strip() for x in line]
        newline[3] = newline[3].replace("mailto:", "")
        output.append(newline)

    return output


def readCsv():
    locations, verticals = readFiles()
    verticals = os.listdir(dir_path + "/files")
    for vertical in verticals:
        found = 0
        while not found:
            csv_choosed = random.choice(
                os.listdir(dir_path + "/files/" + vertical))
            fullpath = dir_path + "/files/%s/%s" % (vertical, csv_choosed)

            with open(fullpath, "r+") as f:
                csv_reader = csv.reader(f, delimiter=",")
                lines = [line[:-1] for line in islice(csv_reader, 0, 15)]

                # Remove last 15 lines from the file
                if lines:
                    new_f = f.readlines()
                    f.seek(0)
                    for line in new_f[15::]:
                        f.write(line)
                    f.truncate()
                    found = 1
                    lines = processLines(lines)
                    newlines = [vertical, *lines]
                    yield newlines
                else:
                    found = 0
